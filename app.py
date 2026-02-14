import io
import os
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'instance', 'onna_business.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'onna-flips-dev-key')

os.makedirs(os.path.join(basedir, 'instance'), exist_ok=True)

db = SQLAlchemy(app)


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date_bought = db.Column(db.Date, nullable=True)
    date_sold = db.Column(db.Date, nullable=True)
    description = db.Column(db.String(200), nullable=False)
    cost = db.Column(db.Float, default=0)
    listing_price = db.Column(db.Float, nullable=True)
    sold_for = db.Column(db.Float, nullable=True)
    status = db.Column(db.String(20), default='Listed')  # Listed, Sold
    notes = db.Column(db.Text, nullable=True)

    @property
    def actual_profit(self):
        if self.sold_for is not None and self.cost is not None:
            return self.sold_for - self.cost
        return None

    @property
    def predicted_profit(self):
        if self.status != 'Sold' and self.listing_price and self.cost is not None:
            return self.listing_price - self.cost
        return None

    @property
    def actual_margin(self):
        if self.sold_for and self.sold_for > 0 and self.cost is not None:
            return round((self.sold_for - self.cost) / self.sold_for, 4)
        return None

    @property
    def days_to_sell(self):
        if self.date_bought and self.date_sold:
            return (self.date_sold - self.date_bought).days
        return None

    @property
    def profit_per_day(self):
        days = self.days_to_sell
        profit = self.actual_profit
        if days and days > 0 and profit is not None:
            return round(profit / days, 2)
        return None

    def to_dict(self):
        return {
            'id': self.id,
            'date_bought': self.date_bought.isoformat() if self.date_bought else None,
            'date_sold': self.date_sold.isoformat() if self.date_sold else None,
            'description': self.description,
            'cost': self.cost,
            'listing_price': self.listing_price,
            'sold_for': self.sold_for,
            'status': self.status,
            'notes': self.notes,
            'actual_profit': self.actual_profit,
            'predicted_profit': self.predicted_profit,
            'actual_margin': self.actual_margin,
            'days_to_sell': self.days_to_sell,
            'profit_per_day': self.profit_per_day,
        }


# --- Page Routes ---

@app.route('/')
def dashboard():
    return render_template('dashboard.html')


@app.route('/inventory')
def inventory():
    return render_template('inventory.html')


@app.route('/add')
def add_item_page():
    return render_template('add_item.html')


@app.route('/edit/<int:item_id>')
def edit_item_page(item_id):
    return render_template('edit_item.html', item_id=item_id)


@app.route('/tax-export')
def tax_export_page():
    return render_template('tax_export.html')


@app.route('/analytics')
def analytics_page():
    return render_template('analytics.html')


# --- API Routes ---

@app.route('/api/items', methods=['GET'])
def get_items():
    status_filter = request.args.get('status')
    query = Item.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    items = query.order_by(Item.date_bought.desc().nullslast()).all()
    return jsonify([item.to_dict() for item in items])


@app.route('/api/items', methods=['POST'])
def create_item():
    data = request.json
    item = Item(
        date_bought=_parse_date(data.get('date_bought')),
        date_sold=_parse_date(data.get('date_sold')),
        description=data['description'],
        cost=float(data.get('cost', 0) or 0),
        listing_price=_parse_float(data.get('listing_price')),
        sold_for=_parse_float(data.get('sold_for')),
        status=data.get('status', 'Listed'),
        notes=data.get('notes'),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/items/<int:item_id>', methods=['GET'])
def get_item(item_id):
    item = Item.query.get_or_404(item_id)
    return jsonify(item.to_dict())


@app.route('/api/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.json
    item.date_bought = _parse_date(data.get('date_bought'))
    item.date_sold = _parse_date(data.get('date_sold'))
    item.description = data['description']
    item.cost = float(data.get('cost', 0) or 0)
    item.listing_price = _parse_float(data.get('listing_price'))
    item.sold_for = _parse_float(data.get('sold_for'))
    item.status = data.get('status', item.status)
    item.notes = data.get('notes')
    db.session.commit()
    return jsonify(item.to_dict())


@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'message': 'Item deleted'})


@app.route('/api/stats', methods=['GET'])
def get_stats():
    all_items = Item.query.all()
    sold_items = [i for i in all_items if i.status == 'Sold']
    listed_items = [i for i in all_items if i.status == 'Listed']

    total_spent = sum(i.cost or 0 for i in all_items)
    total_revenue = sum(i.sold_for or 0 for i in sold_items)
    total_profit = sum(i.actual_profit or 0 for i in sold_items)
    current_invested = sum(i.cost or 0 for i in listed_items)
    predicted_profit = sum(i.predicted_profit or 0 for i in listed_items)

    avg_margin = 0
    margins = [i.actual_margin for i in sold_items if i.actual_margin is not None]
    if margins:
        avg_margin = round(sum(margins) / len(margins), 4)

    avg_days = 0
    days_list = [i.days_to_sell for i in sold_items if i.days_to_sell is not None and i.days_to_sell > 0]
    if days_list:
        avg_days = round(sum(days_list) / len(days_list), 1)

    # Business velocity
    all_bought = [i.date_bought for i in sold_items if i.date_bought]
    all_sold_dates = [i.date_sold for i in sold_items if i.date_sold]
    days_biz = (max(all_sold_dates) - min(all_bought)).days if all_bought and all_sold_dates else 1
    weekly_profit = round(total_profit / max(days_biz / 7, 1))
    money_multiplier = round(total_revenue / total_spent, 1) if total_spent > 0 else 0

    # Monthly profit breakdown
    monthly = {}
    for item in sold_items:
        if item.date_sold and item.actual_profit is not None:
            key = item.date_sold.strftime('%Y-%m')
            monthly[key] = monthly.get(key, 0) + item.actual_profit
    monthly_sorted = sorted(monthly.items())

    # Top items by profit
    top_items = sorted(
        [i for i in sold_items if i.actual_profit is not None],
        key=lambda x: x.actual_profit,
        reverse=True
    )[:10]

    # Top items by profit/day
    top_by_efficiency = sorted(
        [i for i in sold_items if i.profit_per_day is not None and i.profit_per_day > 0],
        key=lambda x: x.profit_per_day,
        reverse=True
    )[:10]

    return jsonify({
        'total_items': len(all_items),
        'sold_count': len(sold_items),
        'listed_count': len(listed_items),
        'total_spent': round(total_spent, 2),
        'total_revenue': round(total_revenue, 2),
        'total_profit': round(total_profit, 2),
        'current_invested': round(current_invested, 2),
        'predicted_profit': round(predicted_profit, 2),
        'avg_margin': avg_margin,
        'avg_days_to_sell': avg_days,
        'weekly_profit': weekly_profit,
        'money_multiplier': money_multiplier,
        'monthly_profit': [{'month': m, 'profit': round(p, 2)} for m, p in monthly_sorted],
        'top_items': [{'description': i.description, 'profit': i.actual_profit} for i in top_items],
        'top_by_efficiency': [
            {'description': i.description, 'profit_per_day': i.profit_per_day, 'days': i.days_to_sell}
            for i in top_by_efficiency
        ],
    })


@app.route('/api/analytics', methods=['GET'])
def get_analytics():
    from collections import defaultdict
    today = date.today()

    all_items = Item.query.all()
    sold = [i for i in all_items if i.status == 'Sold']
    listed = [i for i in all_items if i.status == 'Listed']

    total_cost = sum(i.cost or 0 for i in sold)
    total_revenue = sum(i.sold_for or 0 for i in sold)
    total_profit = sum(i.actual_profit or 0 for i in sold)

    # --- Category Breakdown ---
    categories = defaultdict(list)
    for i in sold:
        desc = i.description.lower()
        if 'table' in desc or 'desk' in desc:
            categories['Tables/Desks'].append(i)
        elif 'chair' in desc or 'stool' in desc:
            categories['Chairs/Seating'].append(i)
        elif 'dresser' in desc or 'night stand' in desc or 'sideboard' in desc:
            categories['Dressers/Storage'].append(i)
        elif 'mirror' in desc or 'picture' in desc:
            categories['Decor/Wall Art'].append(i)
        elif 'rug' in desc:
            categories['Rugs'].append(i)
        elif 'bench' in desc or 'ottoman' in desc:
            categories['Benches/Ottomans'].append(i)
        elif 'hutch' in desc or 'shelf' in desc:
            categories['Hutch/Shelving'].append(i)
        elif 'lamp' in desc:
            categories['Lighting'].append(i)
        else:
            categories['Other'].append(i)

    cat_data = []
    for cat, items_list in sorted(categories.items(), key=lambda x: sum(i.actual_profit or 0 for i in x[1]), reverse=True):
        tp = sum(i.actual_profit or 0 for i in items_list)
        avg_p = tp / len(items_list)
        avg_cost = sum(i.cost or 0 for i in items_list) / len(items_list)
        d = [i.days_to_sell for i in items_list if i.days_to_sell and i.days_to_sell > 0]
        avg_days = round(sum(d) / len(d), 1) if d else 0
        cat_data.append({
            'category': cat, 'count': len(items_list),
            'total_profit': round(tp), 'avg_profit': round(avg_p),
            'avg_cost': round(avg_cost), 'avg_days': avg_days,
        })

    # --- Cost Bracket ROI ---
    cost_brackets_def = [('Free ($0)', 0, 0), ('$1-15', 1, 15), ('$16-30', 16, 30),
                         ('$31-50', 31, 50), ('$51+', 51, 99999)]
    cost_brackets = []
    for label, lo, hi in cost_brackets_def:
        b = [i for i in sold if i.cost is not None and lo <= i.cost <= hi]
        if b:
            tp = sum(i.actual_profit or 0 for i in b)
            roi_items = [i for i in b if i.cost and i.cost > 0 and i.actual_profit is not None]
            avg_roi = round(sum(i.actual_profit / i.cost * 100 for i in roi_items) / len(roi_items)) if roi_items else 0
            cost_brackets.append({
                'bracket': label, 'count': len(b), 'total_profit': round(tp),
                'avg_profit': round(tp / len(b)), 'avg_roi': avg_roi,
            })

    # --- Sell-Through Speed ---
    speed_def = [('Same day', 0, 1), ('Quick (2-7d)', 2, 7), ('Medium (8-14d)', 8, 14),
                 ('Slow (15-30d)', 15, 30), ('Very slow (31+d)', 31, 99999)]
    speed_data = []
    for label, lo, hi in speed_def:
        b = [i for i in sold if i.days_to_sell is not None and lo <= i.days_to_sell <= hi]
        if b:
            tp = sum(i.actual_profit or 0 for i in b)
            speed_data.append({
                'bracket': label, 'count': len(b),
                'total_profit': round(tp), 'avg_profit': round(tp / len(b)),
            })

    # --- Price Bracket (sale price) ---
    price_def = [('$0-50', 0, 50), ('$51-100', 51, 100), ('$101-200', 101, 200),
                 ('$201-300', 201, 300), ('$300+', 301, 99999)]
    price_brackets = []
    for label, lo, hi in price_def:
        b = [i for i in sold if i.sold_for and lo <= i.sold_for <= hi]
        if b:
            tp = sum(i.actual_profit or 0 for i in b)
            d = [i.days_to_sell for i in b if i.days_to_sell and i.days_to_sell > 0]
            price_brackets.append({
                'bracket': label, 'count': len(b),
                'total_profit': round(tp), 'avg_profit': round(tp / len(b)),
                'avg_days': round(sum(d) / len(d), 1) if d else 0,
            })

    # --- Negotiation Analysis ---
    negotiated = [i for i in sold if i.listing_price and i.listing_price > 0 and i.sold_for]
    total_asked = sum(i.listing_price for i in negotiated) if negotiated else 0
    total_got = sum(i.sold_for for i in negotiated) if negotiated else 0
    above = [i for i in negotiated if i.sold_for > i.listing_price]
    at_price = [i for i in negotiated if i.sold_for == i.listing_price]
    below = [i for i in negotiated if i.sold_for < i.listing_price]
    negotiation = {
        'total_items': len(negotiated),
        'total_asked': round(total_asked), 'total_got': round(total_got),
        'avg_discount_pct': round((1 - total_got / total_asked) * 100, 1) if total_asked else 0,
        'above_asking': len(above), 'at_asking': len(at_price), 'below_asking': len(below),
        'above_items': [{'desc': i.description, 'asked': i.listing_price, 'got': i.sold_for,
                         'bonus': round(i.sold_for - i.listing_price)} for i in above],
    }

    # --- Day of Week ---
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_stats = {d: {'count': 0, 'profit': 0} for d in day_names}
    for item in sold:
        if item.date_sold:
            day = day_names[item.date_sold.weekday()]
            day_stats[day]['count'] += 1
            day_stats[day]['profit'] += item.actual_profit or 0
    dow_data = [{'day': d, 'count': s['count'], 'profit': round(s['profit']),
                 'avg_profit': round(s['profit'] / s['count']) if s['count'] else 0}
                for d, s in day_stats.items() if s['count'] > 0]

    # --- Inventory Aging ---
    aging = []
    for i in sorted(listed, key=lambda x: (today - x.date_bought).days if x.date_bought else 0, reverse=True):
        age = (today - i.date_bought).days if i.date_bought else 0
        aging.append({
            'id': i.id, 'description': i.description, 'age_days': age,
            'cost': i.cost, 'listing_price': i.listing_price,
            'predicted_profit': i.predicted_profit,
            'status': 'Stale' if age > 60 else 'Aging' if age > 30 else 'Fresh',
        })

    # --- Best & Worst Flips ---
    sorted_profit = sorted([i for i in sold if i.actual_profit is not None],
                           key=lambda x: x.actual_profit, reverse=True)
    best_flips = [{'desc': i.description, 'cost': i.cost, 'sold': i.sold_for,
                   'profit': i.actual_profit, 'days': i.days_to_sell,
                   'margin': round(i.actual_margin * 100) if i.actual_margin else 0}
                  for i in sorted_profit[:5]]
    worst_flips = [{'desc': i.description, 'cost': i.cost, 'sold': i.sold_for,
                    'profit': i.actual_profit, 'days': i.days_to_sell,
                    'margin': round(i.actual_margin * 100) if i.actual_margin else 0}
                   for i in sorted_profit[-3:]]

    # --- ROI Champions ---
    roi_items = sorted([i for i in sold if i.cost and i.cost > 0 and i.actual_profit is not None],
                       key=lambda x: x.actual_profit / x.cost, reverse=True)
    roi_champs = [{'desc': i.description, 'cost': i.cost, 'sold': i.sold_for,
                   'roi': round(i.actual_profit / i.cost * 100)}
                  for i in roi_items[:7]]

    # --- Speed Demons ---
    fast = sorted([i for i in sold if i.days_to_sell is not None and i.days_to_sell >= 0],
                  key=lambda x: x.days_to_sell)
    speed_demons = [{'desc': i.description, 'days': i.days_to_sell,
                     'profit': i.actual_profit, 'ppd': i.profit_per_day}
                    for i in fast[:7]]

    # --- Business Scorecard ---
    all_bought = [i.date_bought for i in sold if i.date_bought]
    all_sold_dates = [i.date_sold for i in sold if i.date_sold]
    days_biz = (max(all_sold_dates) - min(all_bought)).days if all_bought and all_sold_dates else 1
    margins = [i.actual_margin for i in sold if i.actual_margin is not None]
    avg_margin = round(sum(margins) / len(margins) * 100) if margins else 0
    biggest = max(sold, key=lambda x: x.actual_profit or 0) if sold else None
    fastest_item = min([i for i in sold if i.days_to_sell and i.days_to_sell > 0],
                       key=lambda x: x.days_to_sell, default=None)

    scorecard = {
        'days_in_business': days_biz,
        'weekly_profit': round(total_profit / max(days_biz / 7, 1)),
        'monthly_profit': round(total_profit / max(days_biz / 30, 1)),
        'annualized_profit': round(total_profit / max(days_biz / 365, 1)),
        'avg_flip_cost': round(total_cost / len(sold)) if sold else 0,
        'avg_sale_price': round(total_revenue / len(sold)) if sold else 0,
        'money_multiplier': round(total_revenue / total_cost, 1) if total_cost else 0,
        'profit_per_dollar': round(total_profit / total_cost, 2) if total_cost else 0,
        'items_per_week': round(len(sold) / max(days_biz / 7, 1), 1),
        'avg_margin': avg_margin,
        'biggest_win': f"{biggest.description} (${biggest.actual_profit:.0f})" if biggest else 'N/A',
        'fastest_flip': f"{fastest_item.description} ({fastest_item.days_to_sell}d)" if fastest_item else 'N/A',
        'total_items_sold': len(sold),
        'total_invested': round(total_cost),
        'total_revenue': round(total_revenue),
        'total_profit': round(total_profit),
    }

    return jsonify({
        'categories': cat_data,
        'cost_brackets': cost_brackets,
        'speed_analysis': speed_data,
        'price_brackets': price_brackets,
        'negotiation': negotiation,
        'day_of_week': dow_data,
        'inventory_aging': aging,
        'best_flips': best_flips,
        'worst_flips': worst_flips,
        'roi_champions': roi_champs,
        'speed_demons': speed_demons,
        'scorecard': scorecard,
    })


@app.route('/api/excel-export', methods=['GET'])
def generate_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    include_listed = request.args.get('include_listed') == '1'

    query = Item.query.filter_by(status='Sold')
    if start:
        query = query.filter(Item.date_sold >= start)
    if end:
        query = query.filter(Item.date_sold <= end)
    sold_items = query.order_by(Item.date_sold).all()

    listed_items = []
    if include_listed:
        listed_items = Item.query.filter_by(status='Listed').all()

    wb = Workbook()

    # --- Sold Items Sheet ---
    ws = wb.active
    ws.title = 'Sold Items'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    money_fmt = '$#,##0.00'
    pct_fmt = '0%'
    total_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ['Date Bought', 'Date Sold', 'Item', 'Cost', 'Listing Price', 'Sold For',
               'Profit', 'Margin', 'Days to Sell', 'Profit/Day']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(sold_items, 2):
        ws.cell(row=row_idx, column=1, value=item.date_bought)
        ws.cell(row=row_idx, column=2, value=item.date_sold)
        ws.cell(row=row_idx, column=3, value=item.description)
        ws.cell(row=row_idx, column=4, value=item.cost).number_format = money_fmt
        ws.cell(row=row_idx, column=5, value=item.listing_price).number_format = money_fmt if item.listing_price else 'General'
        ws.cell(row=row_idx, column=6, value=item.sold_for).number_format = money_fmt
        ws.cell(row=row_idx, column=7, value=item.actual_profit).number_format = money_fmt
        ws.cell(row=row_idx, column=8, value=item.actual_margin).number_format = pct_fmt if item.actual_margin else 'General'
        ws.cell(row=row_idx, column=9, value=item.days_to_sell)
        ws.cell(row=row_idx, column=10, value=item.profit_per_day).number_format = money_fmt if item.profit_per_day else 'General'
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Totals row
    tot_row = len(sold_items) + 2
    total_cost = sum(i.cost or 0 for i in sold_items)
    total_revenue = sum(i.sold_for or 0 for i in sold_items)
    total_profit = sum(i.actual_profit or 0 for i in sold_items)
    ws.cell(row=tot_row, column=3, value='TOTALS').font = total_font
    ws.cell(row=tot_row, column=4, value=total_cost).number_format = money_fmt
    ws.cell(row=tot_row, column=6, value=total_revenue).number_format = money_fmt
    ws.cell(row=tot_row, column=7, value=total_profit).number_format = money_fmt
    for col in range(1, 11):
        ws.cell(row=tot_row, column=col).fill = total_fill
        ws.cell(row=tot_row, column=col).font = total_font

    # Auto-width columns
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['C'].width = 35

    # --- Listed Items Sheet ---
    if listed_items:
        ws2 = wb.create_sheet('Unsold Inventory')
        inv_headers = ['Date Bought', 'Item', 'Cost', 'Listing Price', 'Est. Profit']
        for col, h in enumerate(inv_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(listed_items, 2):
            ws2.cell(row=row_idx, column=1, value=item.date_bought)
            ws2.cell(row=row_idx, column=2, value=item.description)
            ws2.cell(row=row_idx, column=3, value=item.cost).number_format = money_fmt
            ws2.cell(row=row_idx, column=4, value=item.listing_price).number_format = money_fmt if item.listing_price else 'General'
            ws2.cell(row=row_idx, column=5, value=item.predicted_profit).number_format = money_fmt if item.predicted_profit else 'General'

        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 14
        ws2.column_dimensions['D'].width = 14
        ws2.column_dimensions['E'].width = 14

    # --- Summary Sheet ---
    ws3 = wb.create_sheet('Summary')
    summary_header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')

    ws3.cell(row=1, column=1, value='Metric').font = header_font
    ws3.cell(row=1, column=1).fill = summary_header_fill
    ws3.cell(row=1, column=2, value='Value').font = header_font
    ws3.cell(row=1, column=2).fill = summary_header_fill

    margins = [i.actual_margin for i in sold_items if i.actual_margin is not None]
    avg_margin = sum(margins) / len(margins) if margins else 0
    days = [i.days_to_sell for i in sold_items if i.days_to_sell and i.days_to_sell > 0]
    avg_days = sum(days) / len(days) if days else 0

    summary = [
        ('Items Sold', len(sold_items)),
        ('Total Cost of Goods', total_cost),
        ('Total Revenue', total_revenue),
        ('Total Profit', total_profit),
        ('Average Margin', avg_margin),
        ('Average Days to Sell', round(avg_days, 1)),
        ('Currently Listed', len(listed_items)),
        ('Current Inventory Cost', sum(i.cost or 0 for i in listed_items)),
        ('Predicted Profit (Listed)', sum(i.predicted_profit or 0 for i in listed_items)),
    ]
    for row_idx, (label, val) in enumerate(summary, 2):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws3.cell(row=row_idx, column=2, value=val)
        if isinstance(val, float) and 'Margin' in label:
            cell.number_format = pct_fmt
        elif isinstance(val, (int, float)) and any(w in label for w in ['Cost', 'Revenue', 'Profit']):
            cell.number_format = money_fmt

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"OnnaFlips_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/tax-export', methods=['GET'])
def generate_tax_pdf():
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    include_listed = request.args.get('include_listed') == '1'

    # Query sold items in date range
    query = Item.query.filter_by(status='Sold')
    if start:
        query = query.filter(Item.date_sold >= start)
    if end:
        query = query.filter(Item.date_sold <= end)
    sold_items = query.order_by(Item.date_sold).all()

    listed_items = []
    if include_listed:
        listed_items = Item.query.filter_by(status='Listed').all()

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, spaceAfter=6)
    elements.append(Paragraph("Onna's Flips - Tax Report", title_style))

    date_range = ""
    if start and end:
        date_range = f"{start.strftime('%m/%d/%Y')} - {end.strftime('%m/%d/%Y')}"
    elif start:
        date_range = f"From {start.strftime('%m/%d/%Y')}"
    elif end:
        date_range = f"Through {end.strftime('%m/%d/%Y')}"
    else:
        date_range = "All dates"
    elements.append(Paragraph(f"Period: {date_range}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%m/%d/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # Summary box
    total_cost = sum(i.cost or 0 for i in sold_items)
    total_revenue = sum(i.sold_for or 0 for i in sold_items)
    total_profit = sum(i.actual_profit or 0 for i in sold_items)

    predicted_profit = sum(i.predicted_profit or 0 for i in listed_items)

    summary_data = [
        ['SUMMARY', '', ''],
        ['Items Sold', 'Total Revenue', 'Total Profit'],
        [str(len(sold_items)), f'${total_revenue:,.2f}', f'${total_profit:,.2f}'],
        ['Cost of Goods', 'Avg Margin', 'Predicted Profit (Listed)'],
    ]

    margins = [i.actual_margin for i in sold_items if i.actual_margin is not None]
    avg_margin = f'{sum(margins)/len(margins)*100:.0f}%' if margins else 'N/A'
    summary_data.append([f'${total_cost:,.2f}', avg_margin, f'${predicted_profit:,.2f}'])

    summary_table = Table(summary_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
    summary_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (2, 0)),
        ('BACKGROUND', (0, 0), (2, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (2, 0), colors.white),
        ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (2, 0), 14),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8f5e9')),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#e8f5e9')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('FONTSIZE', (0, 2), (-1, 2), 14),
        ('FONTSIZE', (0, 4), (-1, 4), 14),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a1a2e')),
        ('GRID', (0, 1), (-1, -1), 0.5, colors.lightgrey),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))

    # Sold items table
    if sold_items:
        elements.append(Paragraph("Sold Items", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))

        table_data = [['Date Sold', 'Date Bought', 'Item', 'Cost', 'Sold For', 'Profit']]
        for item in sold_items:
            table_data.append([
                item.date_sold.strftime('%m/%d/%y') if item.date_sold else '-',
                item.date_bought.strftime('%m/%d/%y') if item.date_bought else '-',
                item.description[:35],
                f'${item.cost:,.2f}' if item.cost else '$0',
                f'${item.sold_for:,.2f}' if item.sold_for else '-',
                f'${item.actual_profit:,.2f}' if item.actual_profit is not None else '-',
            ])

        # Totals row
        table_data.append(['', '', 'TOTALS', f'${total_cost:,.2f}', f'${total_revenue:,.2f}', f'${total_profit:,.2f}'])

        items_table = Table(table_data, colWidths=[0.8*inch, 0.8*inch, 2.3*inch, 0.85*inch, 0.85*inch, 0.85*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(items_table)

    # Monthly Breakout Analysis
    if sold_items:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("Monthly Breakout Analysis", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))

        monthly = {}
        for item in sold_items:
            if item.date_sold:
                key = item.date_sold.strftime('%Y-%m')
                if key not in monthly:
                    monthly[key] = {'items': 0, 'cost': 0, 'revenue': 0, 'profit': 0}
                monthly[key]['items'] += 1
                monthly[key]['cost'] += item.cost or 0
                monthly[key]['revenue'] += item.sold_for or 0
                monthly[key]['profit'] += item.actual_profit or 0

        month_names = {
            '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr', '05': 'May', '06': 'Jun',
            '07': 'Jul', '08': 'Aug', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec',
        }

        month_data = [['Month', 'Items Sold', 'Cost', 'Revenue', 'Profit', 'Margin']]
        for key in sorted(monthly.keys()):
            m = monthly[key]
            y, mo = key.split('-')
            margin = f'{m["profit"]/m["revenue"]*100:.0f}%' if m['revenue'] else 'N/A'
            month_data.append([
                f'{month_names[mo]} {y}',
                str(m['items']),
                f'${m["cost"]:,.2f}',
                f'${m["revenue"]:,.2f}',
                f'${m["profit"]:,.2f}',
                margin,
            ])
        # Totals
        month_data.append([
            'TOTAL', str(len(sold_items)),
            f'${total_cost:,.2f}', f'${total_revenue:,.2f}', f'${total_profit:,.2f}',
            f'{total_profit/total_revenue*100:.0f}%' if total_revenue else 'N/A',
        ])

        month_table = Table(month_data, colWidths=[1.1*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
        month_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(month_table)

    # Listed items (unsold inventory)
    if listed_items:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("Unsold Inventory", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))

        inv_data = [['Date Bought', 'Item', 'Cost', 'Listing Price', 'Est. Profit']]
        inv_total_cost = 0
        for item in listed_items:
            inv_total_cost += item.cost or 0
            inv_data.append([
                item.date_bought.strftime('%m/%d/%y') if item.date_bought else '-',
                item.description[:40],
                f'${item.cost:,.2f}' if item.cost else '$0',
                f'${item.listing_price:,.2f}' if item.listing_price else '-',
                f'${item.predicted_profit:,.2f}' if item.predicted_profit else '-',
            ])
        inv_data.append(['', 'TOTAL INVESTED', f'${inv_total_cost:,.2f}', '', ''])

        inv_table = Table(inv_data, colWidths=[0.9*inch, 2.5*inch, 0.9*inch, 1*inch, 0.9*inch])
        inv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6c757d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3cd')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(inv_table)

    doc.build(elements)
    buf.seek(0)

    filename = f"OnnaFlips_TaxReport_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _parse_float(val):
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
